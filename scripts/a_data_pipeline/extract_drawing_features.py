"""
笔迹特征提取脚本
从 MMSE_MoCA XLSX 文件的「绘图数据」sheet 中提取笔迹特征。

输出:
  - D:\vscode\2\output3\drawing_features.csv     (每个被试一行，含所有绘图特征)
  - D:\vscode\2\output3\drawing_features_detail.csv (每个被试每任务一行，详细特征)
"""

import os
import json
import math
import numpy as np
import openpyxl
import csv
from collections import defaultdict

INPUT_DIR = r"D:\vscode\2\unused\MMSE_MoCA"
OUTPUT_DIR = r"D:\vscode\2\output3"

# ─── 工具函数 ───────────────────────────────────────────

def concat_row_json(ws, row_idx):
    """拼接一行中从 E 列开始所有分段的 JSON 字符串"""
    parts = []
    for col in range(5, ws.max_column + 1):  # 5 = E 列
        cell = ws.cell(row=row_idx, column=col)
        if cell.value is not None:
            parts.append(str(cell.value))
    return "".join(parts)


def parse_strokes(full_text):
    """解析拼接后的 JSON 字符串"""
    text = full_text.strip()
    if text in ("", "[]"):
        return None
    return json.loads(text)


def extract_all_points(strokes):
    """从 strokes 中提取所有点信息"""
    points = []  # [(x, y, timestamp, linewidth), ...]
    for stroke in strokes:
        for pt in stroke:
            points.append((
                pt.get("x", 0.0),
                pt.get("y", 0.0),
                pt.get("timeStamp", 0.0),
                pt.get("lineWidth", 1.0),
            ))
    return points


def fast_point_list(strokes):
    """返回 numpy 数组: (N, 4) = x, y, timestamp, linewidth"""
    data = []
    for stroke in strokes:
        for pt in stroke:
            data.append([
                pt.get("x", 0.0),
                pt.get("y", 0.0),
                pt.get("timeStamp", 0.0),
                pt.get("lineWidth", 1.0),
            ])
    if not data:
        return np.zeros((0, 4))
    return np.array(data)


# ─── 通用特征计算 ─────────────────────────────────────

def compute_common_features(strokes):
    """
    计算 16 个通用特征。
    输入: strokes (list of list of dict)
    返回: dict of feature_name -> value
    """
    pts = fast_point_list(strokes)
    if len(pts) == 0:
        return {f: np.nan for f in COMMON_FEATURE_NAMES}

    x, y, t, lw = pts[:, 0], pts[:, 1], pts[:, 2], pts[:, 3]
    feats = {}

    # 1. point_count
    feats["point_count"] = len(pts)

    # 2. stroke_count
    feats["stroke_count"] = len(strokes)

    # 3. duration (seconds)
    t_span = np.max(t) - np.min(t)
    feats["duration"] = t_span if t_span > 0 else 0.0

    # 4. path_length
    path_len = 0.0
    for stroke in strokes:
        for i in range(1, len(stroke)):
            dx = stroke[i]["x"] - stroke[i-1]["x"]
            dy = stroke[i]["y"] - stroke[i-1]["y"]
            path_len += math.sqrt(dx*dx + dy*dy)
    feats["path_length"] = path_len

    # 5-8. bounding box
    x_min, x_max = np.min(x), np.max(x)
    y_min, y_max = np.min(y), np.max(y)
    bbox_w = x_max - x_min
    bbox_h = y_max - y_min
    feats["bbox_width"] = bbox_w
    feats["bbox_height"] = bbox_h
    feats["bbox_area"] = bbox_w * bbox_h
    feats["bbox_aspect_ratio"] = max(bbox_w, bbox_h) / min(bbox_w, bbox_h) if min(bbox_w, bbox_h) > 0 else 1.0

    # 9-10. speed
    speeds = []
    for stroke in strokes:
        for i in range(1, len(stroke)):
            dx = stroke[i]["x"] - stroke[i-1]["x"]
            dy = stroke[i]["y"] - stroke[i-1]["y"]
            dt = stroke[i]["timeStamp"] - stroke[i-1]["timeStamp"]
            if dt > 0.001:
                speeds.append(math.sqrt(dx*dx + dy*dy) / dt)
    if speeds:
        feats["speed_mean"] = np.mean(speeds)
        feats["speed_std"] = np.std(speeds)
    else:
        feats["speed_mean"] = 0.0
        feats["speed_std"] = 0.0

    # 11-13. pause (time gap > 1.0 second between consecutive points)
    pause_threshold = 1.0
    pause_count = 0
    pause_total = 0.0
    for stroke in strokes:
        for i in range(1, len(stroke)):
            dt = stroke[i]["timeStamp"] - stroke[i-1]["timeStamp"]
            if dt > pause_threshold:
                pause_count += 1
                pause_total += (dt - pause_threshold)
    feats["pause_count"] = pause_count
    feats["pause_total_duration"] = pause_total
    feats["pause_ratio"] = pause_total / t_span if t_span > 0 else 0.0

    # 14-15. line_width
    feats["line_width_mean"] = float(np.mean(lw))
    feats["line_width_std"] = float(np.std(lw))

    # 16. grid_occupancy_ratio
    grid_size = 20
    if bbox_w > 0 and bbox_h > 0:
        x_grid = np.clip(((x - x_min) / bbox_w * grid_size).astype(int), 0, grid_size - 1)
        y_grid = np.clip(((y - y_min) / bbox_h * grid_size).astype(int), 0, grid_size - 1)
        occupied = set(zip(x_grid, y_grid))
        feats["grid_occupancy_ratio"] = len(occupied) / (grid_size * grid_size)
    else:
        feats["grid_occupancy_ratio"] = 0.0

    # 17. spatial_entropy
    if bbox_w > 0 and bbox_h > 0:
        hist, _, _ = np.histogram2d(x, y, bins=grid_size,
                                     range=[[x_min, x_max], [y_min, y_max]])
        hist = hist.flatten()
        hist = hist / np.sum(hist)
        hist = hist[hist > 0]
        feats["spatial_entropy"] = float(-np.sum(hist * np.log(hist + 1e-12)))
    else:
        feats["spatial_entropy"] = 0.0

    return feats


COMMON_FEATURE_NAMES = [
    "point_count", "stroke_count", "duration", "path_length",
    "bbox_width", "bbox_height", "bbox_area", "bbox_aspect_ratio",
    "speed_mean", "speed_std",
    "pause_count", "pause_total_duration", "pause_ratio",
    "line_width_mean", "line_width_std",
    "grid_occupancy_ratio", "spatial_entropy",
]


# ─── 辅助几何函数 ──────────────────────────────────────

def compute_direction_angles(strokes):
    """计算每个连续点对的方向角（弧度）"""
    angles = []
    for stroke in strokes:
        for i in range(1, len(stroke)):
            dx = stroke[i]["x"] - stroke[i-1]["x"]
            dy = stroke[i]["y"] - stroke[i-1]["y"]
            if abs(dx) > 0.001 or abs(dy) > 0.001:
                angles.append(math.atan2(dy, dx))
    return angles


def line_segment_intersection(p1, p2, p3, p4):
    """检查线段 p1-p2 和 p3-p4 是否相交（不含端点接触）"""
    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])
    def on_segment(p, q, r):
        return (min(p[0], r[0]) <= q[0] <= max(p[0], r[0]) and
                min(p[1], r[1]) <= q[1] <= max(p[1], r[1]))

    d1 = cross(p3, p4, p1)
    d2 = cross(p3, p4, p2)
    d3 = cross(p1, p2, p3)
    d4 = cross(p1, p2, p4)

    if ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
       ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0)):
        return True
    if abs(d1) < 1e-9 and on_segment(p3, p1, p4): return True
    if abs(d2) < 1e-9 and on_segment(p3, p2, p4): return True
    if abs(d3) < 1e-9 and on_segment(p1, p3, p2): return True
    if abs(d4) < 1e-9 and on_segment(p1, p4, p2): return True
    return False


def count_intersections(strokes, sample_step=3):
    """简化的线段交叉计数：对每个 stroke 取采样点，检测与其他 stroke 的交叉"""
    # 构建所有 stroke 的线段列表
    all_segments = []
    for si, stroke in enumerate(strokes):
        for i in range(0, len(stroke) - 1, sample_step):
            j = min(i + sample_step, len(stroke) - 1)
            all_segments.append((si, stroke[i], stroke[j]))

    count = 0
    n = len(all_segments)
    for i in range(n):
        for j in range(i + 1, n):
            if all_segments[i][0] == all_segments[j][0]:
                continue  # 同一 stroke 内的不检测
            p1 = (all_segments[i][1]["x"], all_segments[i][1]["y"])
            p2 = (all_segments[i][2]["x"], all_segments[i][2]["y"])
            p3 = (all_segments[j][1]["x"], all_segments[j][1]["y"])
            p4 = (all_segments[j][2]["x"], all_segments[j][2]["y"])
            if line_segment_intersection(p1, p2, p3, p4):
                count += 1
    return count


def count_direction_changes(angles, threshold_deg=45):
    """计算方向变化超过阈值的次数"""
    if len(angles) < 2:
        return 0
    count = 0
    for i in range(1, len(angles)):
        diff = abs(angles[i] - angles[i-1])
        if diff > math.pi:
            diff = 2 * math.pi - diff
        if math.degrees(diff) > threshold_deg:
            count += 1
    return count


def prune_straight_points(stroke, tolerance=0.5):
    """
    移除直线上的冗余中间点（Ramer-Douglas-Peucker 简化）。
    这里用一个简化的方法：只保留方向变化明显的点。
    返回简化的点列表。
    """
    if len(stroke) < 3:
        return stroke
    # 简化版：保留拐点
    kept = [stroke[0]]
    for i in range(1, len(stroke) - 1):
        prev_angle = math.atan2(
            stroke[i]["y"] - stroke[i-1]["y"],
            stroke[i]["x"] - stroke[i-1]["x"]
        )
        next_angle = math.atan2(
            stroke[i+1]["y"] - stroke[i]["y"],
            stroke[i+1]["x"] - stroke[i]["x"]
        )
        diff = abs(next_angle - prev_angle)
        if diff > math.pi:
            diff = 2 * math.pi - diff
        if math.degrees(diff) > 15:  # 角度变化 > 15度就保留
            kept.append(stroke[i])
    kept.append(stroke[-1])
    return kept


# ─── MMSE 图形复制特征 ──────────────────────────────

def compute_mmse_shape_features(strokes, common_feats):
    """MMSE 图形复制专属特征"""
    feats = {}

    # closed_loop_count
    closed = 0
    for stroke in strokes:
        if len(stroke) >= 3:
            dx = stroke[-1]["x"] - stroke[0]["x"]
            dy = stroke[-1]["y"] - stroke[0]["y"]
            start_end_dist = math.sqrt(dx*dx + dy*dy)
            # 计算 stroke 的路径长度
            stroke_len = 0.0
            for i in range(1, len(stroke)):
                sx = stroke[i]["x"] - stroke[i-1]["x"]
                sy = stroke[i]["y"] - stroke[i-1]["y"]
                stroke_len += math.sqrt(sx*sx + sy*sy)
            if stroke_len > 0 and start_end_dist < 0.15 * stroke_len:
                closed += 1
    feats["closed_loop_count"] = closed

    # corner_count, sharp_turn_count
    angles = compute_direction_angles(strokes)
    corner_count = 0
    sharp_count = 0
    for i in range(1, len(angles)):
        diff = abs(angles[i] - angles[i-1])
        if diff > math.pi:
            diff = 2 * math.pi - diff
        deg = math.degrees(diff)
        if deg > 60:
            corner_count += 1
        if deg > 120:
            sharp_count += 1
    feats["corner_count"] = corner_count
    feats["sharp_turn_count"] = sharp_count

    # intersection_count
    feats["intersection_count"] = count_intersections(strokes)

    # symmetry_score (垂直轴对称)
    pts = fast_point_list(strokes)
    if len(pts) > 0:
        x = pts[:, 0]
        x_mid = (np.min(x) + np.max(x)) / 2.0
        # 将左半部分的点镜像到右半，比较分布
        left = x[x < x_mid]
        right = x[x > x_mid]
        right_mirror = 2 * x_mid - right
        all_x_mirrored = np.concatenate([left, right_mirror])
        if len(all_x_mirrored) > 1 and len(x) > 1:
            # 用 K-S 统计量或简单直方图对比
            hist_full, _ = np.histogram(x, bins=20, range=(np.min(x), np.max(x)))
            hist_sym, _ = np.histogram(all_x_mirrored, bins=20, range=(np.min(x), np.max(x)))
            hist_full = hist_full / np.sum(hist_full)
            hist_sym = hist_sym / np.sum(hist_sym)
            # Bhattacharyya coefficient (相似度)
            bc = np.sum(np.sqrt(hist_full * hist_sym + 1e-12))
            feats["symmetry_score"] = float(bc)
        else:
            feats["symmetry_score"] = np.nan
    else:
        feats["symmetry_score"] = np.nan

    return feats


# ─── MoCA 连线测验特征 ──────────────────────────────

def compute_moca_trail_features(strokes, common_feats):
    """MoCA 连线测验专属特征"""
    feats = {}
    pts = fast_point_list(strokes)

    # path_efficiency = 直线距离 / 总路径
    if common_feats["path_length"] > 0 and len(pts) > 1:
        start = pts[0, :2]
        end = pts[-1, :2]
        straight_dist = math.sqrt((end[0]-start[0])**2 + (end[1]-start[1])**2)
        feats["path_efficiency"] = straight_dist / common_feats["path_length"]
    else:
        feats["path_efficiency"] = np.nan

    # direction_change_count
    angles = compute_direction_angles(strokes)
    feats["direction_change_count"] = count_direction_changes(angles, 45)

    # backtrack_ratio
    backtrack_dist = 0.0
    if len(pts) > 1:
        overall_dir = pts[-1, :2] - pts[0, :2]
        overall_dir_norm = np.linalg.norm(overall_dir)
        if overall_dir_norm > 1e-6:
            overall_dir = overall_dir / overall_dir_norm
            for i in range(len(pts) - 1):
                seg = pts[i+1, :2] - pts[i, :2]
                seg_len = np.linalg.norm(seg)
                if seg_len > 0:
                    proj = np.dot(seg / seg_len, overall_dir)
                    if proj < -0.3:  # 明显反向
                        backtrack_dist += seg_len
        total_dist = common_feats["path_length"]
        feats["backtrack_ratio"] = backtrack_dist / total_dist if total_dist > 0 else 0.0
    else:
        feats["backtrack_ratio"] = 0.0

    # segment_count (每个方向一致的连续段)
    segments = 0
    for stroke in strokes:
        pruned = prune_straight_points(stroke)
        segments += max(1, len(pruned) - 1)
    feats["segment_count"] = segments

    # crossing_count
    feats["crossing_count"] = count_intersections(strokes)

    return feats


# ─── MoCA 复制立方体特征 ────────────────────────────

def compute_moca_cube_features(strokes, common_feats):
    """MoCA 复制立方体专属特征"""
    feats = {}

    # corner_count
    angles = compute_direction_angles(strokes)
    corner_count = 0
    for i in range(1, len(angles)):
        diff = abs(angles[i] - angles[i-1])
        if diff > math.pi:
            diff = 2 * math.pi - diff
        if math.degrees(diff) > 60:
            corner_count += 1
    feats["corner_count"] = corner_count

    # line_segment_count: 简化后的直线段数
    segments = 0
    for stroke in strokes:
        pruned = prune_straight_points(stroke)
        segments += max(1, len(pruned) - 1)
    feats["line_segment_count"] = segments

    # parallel_score
    # 提取主要的线段方向，计算平行度
    directions = []
    for stroke in strokes:
        for i in range(0, max(1, len(stroke) - 5), 5):
            j = min(i + 5, len(stroke) - 1)
            dx = stroke[j]["x"] - stroke[i]["x"]
            dy = stroke[j]["y"] - stroke[i]["y"]
            if abs(dx) > 0.5 or abs(dy) > 0.5:
                angle = math.atan2(dy, dx) % math.pi  # 归一化到 [0, pi)
                directions.append(angle)
    if len(directions) >= 2:
        # 聚类方向，检查是否有成对平行（差值 < 15度）
        dirs = np.array(directions)
        # 计算成对差值
        parallel_pairs = 0
        total_pairs = 0
        for i in range(min(len(dirs), 30)):
            for j in range(i+1, min(len(dirs), 30)):
                diff = abs(dirs[i] - dirs[j])
                if diff > math.pi/2:
                    diff = math.pi - diff
                if math.degrees(diff) < 15:
                    parallel_pairs += 1
                total_pairs += 1
        feats["parallel_score"] = parallel_pairs / total_pairs if total_pairs > 0 else 0.0
    else:
        feats["parallel_score"] = 0.0

    # angle_consistency (角度一致性：std of direction angles)
    if len(directions) >= 2:
        feats["angle_consistency"] = float(np.std(directions))
    else:
        feats["angle_consistency"] = np.nan

    # intersection_count
    feats["intersection_count"] = count_intersections(strokes)

    # symmetry_score
    pts = fast_point_list(strokes)
    if len(pts) > 0:
        x = pts[:, 0]
        x_mid = (np.min(x) + np.max(x)) / 2.0
        left = x[x < x_mid]
        right = x[x > x_mid]
        right_mirror = 2 * x_mid - right
        all_x_mirrored = np.concatenate([left, right_mirror])
        if len(all_x_mirrored) > 1 and len(x) > 1:
            hist_full, _ = np.histogram(x, bins=20, range=(np.min(x), np.max(x)))
            hist_sym, _ = np.histogram(all_x_mirrored, bins=20, range=(np.min(x), np.max(x)))
            hist_full = hist_full / (np.sum(hist_full) + 1e-12)
            hist_sym = hist_sym / (np.sum(hist_sym) + 1e-12)
            feats["symmetry_score"] = float(np.sum(np.sqrt(hist_full * hist_sym + 1e-12)))
        else:
            feats["symmetry_score"] = np.nan
    else:
        feats["symmetry_score"] = np.nan

    return feats


# ─── MoCA 绘制钟表特征 ──────────────────────────────

def compute_moca_clock_features(strokes, common_feats):
    """MoCA 绘制钟表专属特征"""
    feats = {}
    pts = fast_point_list(strokes)

    if len(pts) == 0:
        for f in CLOCK_FEATURE_NAMES:
            feats[f] = np.nan
        return feats

    x, y = pts[:, 0], pts[:, 1]
    x_mid = (np.min(x) + np.max(x)) / 2.0
    y_mid = (np.min(y) + np.max(y)) / 2.0

    # 计算各点到 bbox 中心 (x_mid, y_mid) 的距离
    radii = np.sqrt((x - x_mid)**2 + (y - y_mid)**2)
    angles_from_center = np.arctan2(y - y_mid, x - x_mid)

    # circle_roundness: 半径的变异系数
    r_mean = np.mean(radii)
    r_std = np.std(radii)
    feats["circle_roundness"] = float(1.0 / (1.0 + r_std / r_mean)) if r_mean > 0 else 0.0

    # circle_closure_error: 角度覆盖的缺失
    # 将角度排序，找最大间隙
    sorted_angles = np.sort(angles_from_center)
    max_gap = 0.0
    for i in range(len(sorted_angles) - 1):
        gap = sorted_angles[i+1] - sorted_angles[i]
        if gap > max_gap:
            max_gap = gap
    total_gap = max_gap
    feats["circle_closure_error"] = float(total_gap)

    # center_offset: 半径的偏度
    feats["center_offset"] = float(np.std(radii) / r_mean) if r_mean > 0 else 0.0

    # number_region_coverage: 12 个角扇区的覆盖情况
    sector_angles = np.floor((angles_from_center + math.pi) / (2 * math.pi) * 12).astype(int)
    sector_angles = np.clip(sector_angles, 0, 11)
    covered_sectors = len(set(sector_angles))
    feats["number_region_coverage"] = covered_sectors / 12.0

    # number_distribution_balance: 各扇区点数分布的均匀性
    sector_counts = np.bincount(sector_angles, minlength=12)
    if np.sum(sector_counts) > 0:
        sector_probs = sector_counts / np.sum(sector_counts)
        sector_probs = sector_probs[sector_probs > 0]
        entropy = -np.sum(sector_probs * np.log(sector_probs + 1e-12))
        max_entropy = np.log(12)
        feats["number_distribution_balance"] = float(entropy / max_entropy) if max_entropy > 0 else 0.0
    else:
        feats["number_distribution_balance"] = 0.0

    # radial_distribution_entropy: 半径分布的熵
    if len(radii) > 0 and r_mean > 0:
        r_bins = 10
        r_hist, _ = np.histogram(radii, bins=r_bins, range=(0, r_mean * 2))
        r_hist = r_hist / np.sum(r_hist)
        r_hist = r_hist[r_hist > 0]
        feats["radial_distribution_entropy"] = float(-np.sum(r_hist * np.log(r_hist + 1e-12)))
    else:
        feats["radial_distribution_entropy"] = 0.0

    # hand_count_estimate: 估算指针数量
    # 找出从中心出发的长直线
    center = np.array([x_mid, y_mid])
    long_strokes = 0
    stroke_info = []
    for stroke in strokes:
        if len(stroke) < 2:
            continue
        start = np.array([stroke[0]["x"], stroke[0]["y"]])
        end = np.array([stroke[-1]["x"], stroke[-1]["y"]])
        start_d = np.linalg.norm(start - center)
        end_d = np.linalg.norm(end - center)
        stroke_path = 0.0
        for i in range(1, len(stroke)):
            sx = stroke[i]["x"] - stroke[i-1]["x"]
            sy = stroke[i]["y"] - stroke[i-1]["y"]
            stroke_path += math.sqrt(sx*sx + sy*sy)
        # 如果一端接近中心且路径较长，可能是指针
        min_d = min(start_d, end_d)
        if min_d < max(bbox_w := (np.max(x) - np.min(x)), np.max(y) - np.min(y)) * 0.3 and stroke_path > bbox_w * 0.2:
            long_strokes += 1
            stroke_info.append({
                "length": stroke_path,
                "angle": math.atan2(end[1] - start[1], end[0] - start[0])
            })
    feats["hand_count_estimate"] = long_strokes

    # hand_angle_feature: 两个最长指针之间的角度
    if len(stroke_info) >= 2:
        sorted_strokes = sorted(stroke_info, key=lambda s: s["length"], reverse=True)
        ang1 = sorted_strokes[0]["angle"]
        ang2 = sorted_strokes[1]["angle"]
        diff = abs(ang1 - ang2)
        if diff > math.pi:
            diff = 2 * math.pi - diff
        feats["hand_angle_feature"] = float(math.degrees(diff))
    else:
        feats["hand_angle_feature"] = np.nan

    return feats


CLOCK_FEATURE_NAMES = [
    "circle_roundness", "circle_closure_error", "center_offset",
    "number_region_coverage", "number_distribution_balance",
    "radial_distribution_entropy", "hand_count_estimate", "hand_angle_feature",
]


# ─── 任务定义 ──────────────────────────────────────────

TASK_DEFS = [
    {
        "prefix": "drawing_mmse_shape",
        "scale": "MMSE",
        "section": "5",
        "question": "1",
        "description": "MMSE图形复制",
        "extra_features_fn": compute_mmse_shape_features,
        "extra_feature_names": [
            "closed_loop_count", "corner_count", "sharp_turn_count",
            "intersection_count", "symmetry_score",
        ],
    },
    {
        "prefix": "drawing_moca_trail",
        "scale": "MOCA",
        "section": "1",
        "question": "1",
        "description": "MoCA连线测验",
        "extra_features_fn": compute_moca_trail_features,
        "extra_feature_names": [
            "path_efficiency", "direction_change_count", "backtrack_ratio",
            "segment_count", "crossing_count",
        ],
    },
    {
        "prefix": "drawing_moca_cube",
        "scale": "MOCA",
        "section": "1",
        "question": "2",
        "description": "MoCA复制立方体",
        "extra_features_fn": compute_moca_cube_features,
        "extra_feature_names": [
            "corner_count", "line_segment_count", "parallel_score",
            "angle_consistency", "intersection_count", "symmetry_score",
        ],
    },
    {
        "prefix": "drawing_moca_clock",
        "scale": "MOCA",
        "section": "1",
        "question": "3",
        "description": "MoCA绘制钟表",
        "extra_features_fn": compute_moca_clock_features,
        "extra_feature_names": CLOCK_FEATURE_NAMES,
    },
]


# ─── 主流程 ─────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    files = sorted([
        f for f in os.listdir(INPUT_DIR)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ])

    # 收集所有被试的特征
    all_subject_features = []  # 每项: dict(subject_id, ...features...)
    detail_rows = []           # 详细数据

    for fname in files:
        path = os.path.join(INPUT_DIR, fname)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["绘图数据"]
        raw_id = str(ws["A1"].value).strip()
        # 标准化 subject_id
        subject_id = raw_id.replace("_____ATHENA", "").replace("_____", "").strip()
        # 确保是 ATH 开头的数字编号
        if subject_id.startswith("ATHENA"):
            subject_id = subject_id.replace("ATHENA", "ATH")

        row_data = {}  # subject_id -> combined features

        for task in TASK_DEFS:
            prefix = task["prefix"]
            scale = task["scale"]
            section = task["section"]
            question = task["question"]

            # 找到对应的行
            strokes = None
            for row_idx in range(3, 7):
                row_scale = str(ws[f"B{row_idx}"].value or "").strip()
                row_section = str(ws[f"C{row_idx}"].value or "").strip()
                row_question = str(ws[f"D{row_idx}"].value or "").strip()

                if row_scale == scale and row_section == section and row_question == question:
                    full_text = concat_row_json(ws, row_idx)
                    strokes = parse_strokes(full_text)
                    break

            if strokes is None:
                # 没有找到数据
                detail_rows.append({
                    "subject_id": subject_id,
                    "task": prefix,
                    "has_data": False,
                })
                # 填充 NaN
                for fn in COMMON_FEATURE_NAMES + task["extra_feature_names"]:
                    row_data[f"{prefix}_{fn}"] = np.nan
                continue

            # 计算通用特征
            common_feats = compute_common_features(strokes)
            for fn in COMMON_FEATURE_NAMES:
                row_data[f"{prefix}_{fn}"] = common_feats.get(fn, np.nan)

            # 计算专属特征
            extra_feats = task["extra_features_fn"](strokes, common_feats)
            for fn in task["extra_feature_names"]:
                row_data[f"{prefix}_{fn}"] = extra_feats.get(fn, np.nan)

            # 记录详细数据
            detail_rows.append({
                "subject_id": subject_id,
                "task": prefix,
                "has_data": True,
                **{f"{prefix}_{fn}": common_feats.get(fn, np.nan) for fn in COMMON_FEATURE_NAMES},
                **{f"{prefix}_{fn}": extra_feats.get(fn, np.nan) for fn in task["extra_feature_names"]},
            })

        row_data["subject_id"] = subject_id
        all_subject_features.append(row_data)
        print(f"  {subject_id}: {len(row_data)-1} features extracted")

    # ── 输出 CSV ──

    # 1. 每被试一行 (宽表)
    all_feature_keys = set()
    for row in all_subject_features:
        all_feature_keys.update(row.keys())
    all_feature_keys.discard("subject_id")
    # 排序: 按前缀
    def sort_key(k):
        prefixes = ["drawing_mmse_shape", "drawing_moca_trail", "drawing_moca_cube", "drawing_moca_clock"]
        for i, p in enumerate(prefixes):
            if k.startswith(p):
                return (i, k)
        return (99, k)
    sorted_keys = sorted(all_feature_keys, key=sort_key)

    output_csv = os.path.join(OUTPUT_DIR, "drawing_features.csv")
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["subject_id"] + sorted_keys)
        writer.writeheader()
        for row in all_subject_features:
            writer.writerow({k: row.get(k, "") for k in ["subject_id"] + sorted_keys})
    print(f"\n=> {output_csv} ({len(all_subject_features)} subjects, {len(sorted_keys)} features)")

    # 2. 详细数据 (每被试每任务一行)
    detail_csv = os.path.join(OUTPUT_DIR, "drawing_features_detail.csv")
    if detail_rows:
        detail_keys = set()
        for row in detail_rows:
            detail_keys.update(row.keys())
        detail_keys.discard("subject_id")
        detail_keys.discard("task")
        detail_keys.discard("has_data")
        sorted_detail = sorted(detail_keys, key=sort_key)

        with open(detail_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["subject_id", "task", "has_data"] + sorted_detail)
            writer.writeheader()
            for row in detail_rows:
                writer.writerow({k: row.get(k, "") for k in ["subject_id", "task", "has_data"] + sorted_detail})
        print(f"=> {detail_csv} ({len(detail_rows)} rows)")

    # 3. 统计摘要
    stats_path = os.path.join(OUTPUT_DIR, "drawing_features_summary.txt")
    with open(stats_path, "w", encoding="utf-8") as f:
        f.write(f"Drawing Feature Extraction Summary\n")
        f.write(f"{'='*60}\n")
        f.write(f"Subjects: {len(all_subject_features)}\n")
        f.write(f"Total features: {len(sorted_keys)}\n\n")
        for task in TASK_DEFS:
            task_keys = [k for k in sorted_keys if k.startswith(task["prefix"])]
            has_data_count = sum(1 for row in detail_rows
                                 if row["task"] == task["prefix"] and row.get("has_data"))
            f.write(f"{task['description']} ({task['prefix']}):\n")
            f.write(f"  Subjects with data: {has_data_count}\n")
            f.write(f"  Features: {len(task_keys)}\n")
            f.write(f"  Common features: {len(COMMON_FEATURE_NAMES)}\n")
            f.write(f"  Task-specific features: {len(task['extra_feature_names'])}\n\n")
    print(f"=> {stats_path}")

    print("\nDone! Feature extraction complete.")


if __name__ == "__main__":
    main()
