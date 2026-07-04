"""
完整的 merged_dataset.csv 生成脚本

数据源:
  1. VR 日志: unused/exe_release_20260415/log/
     每个被试一个文件夹，内含四宫格/九宫格日志。
     任务类型通过 INIT_EXP 中的 COMM.NAME 识别:
       "田字格社区" → grid4
       "九宫格社区" → grid9
     _log_N 后缀表示续存日志，同一任务类型合并处理。

  2. 量表 Excel: unused/MMSE_MoCA/*.xlsx
     从「量表总分」sheet 提取 MMSE, MOCA, CDR, HIS。

  3. 笔迹特征: output3/drawing_features.csv

输出到 output1/:
  - merged_dataset.csv  (总表)
  - scale_scores.csv     (量表中间表)
  - vr_features.csv      (VR 特征中间表)
"""

import os
import re
import json
import csv
import math
import glob
import numpy as np
import openpyxl
from collections import defaultdict

BASE_DIR = r"D:\vscode\2"
VR_LOG_DIR = os.path.join(BASE_DIR, "unused", "exe_release_20260415", "log")
SCALE_DIR = os.path.join(BASE_DIR, "unused", "MMSE_MoCA")
DRAWING_CSV = os.path.join(BASE_DIR, "output3", "drawing_features.csv")
OUTPUT_DIR = os.path.join(BASE_DIR, "output1")

# ─── VR 日志解析 ───────────────────────────────────────

TASK_NAME_MAP = {
    "田字格社区": "grid4",
    "九宫格社区": "grid9",
}

# 文件名字面匹配（用于日志中检测不到 INIT_EXP 时的回退）
FILENAME_TASK_MAP = {
    "四宫格": "grid4",
    "田字格": "grid4",
    "九宫格": "grid9",
}


def detect_task_type_from_filename(fname):
    """从文件名检测任务类型"""
    for keyword, tt in FILENAME_TASK_MAP.items():
        if keyword in fname:
            return tt
    return None


def read_init_exp_task_type(fpath):
    """读取日志文件前 100 行，查找 INIT_EXP 中的 COMM.NAME 来确定任务类型"""
    try:
        with open(fpath, "r", encoding="utf-8", errors="replace") as f:
            for i, line in enumerate(f):
                if i > 100:
                    break
                if "INIT_EXP" not in line:
                    continue
                try:
                    parts = line.strip().split("|", 3)
                    if len(parts) >= 4:
                        payload = json.loads(parts[3])
                        comm = payload.get("COMM", {})
                        name = comm.get("NAME", "")
                        if name in TASK_NAME_MAP:
                            return TASK_NAME_MAP[name]
                except (json.JSONDecodeError, KeyError):
                    continue
    except Exception:
        pass
    return None


def parse_log_events(fpaths):
    """从一组日志文件中解析所有事件和 BASE_INFO。
    返回: {events: [...], base_infos: [...], task_type: str}
    """
    events = []
    base_infos = []
    task_type = None
    v_stop_thr = 0.01

    for fpath in sorted(fpaths):
        # 从文件名尝试检测任务类型
        fn_task = detect_task_type_from_filename(os.path.basename(fpath))

        try:
            with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    try:
                        parts = line.strip().split("|", 3)
                        if len(parts) < 4:
                            continue
                        rel_time = float(parts[0])
                        real_time = parts[1]
                        event_type = parts[2].strip()
                        payload = json.loads(parts[3])
                    except (json.JSONDecodeError, ValueError):
                        continue

                    # 任务类型检测
                    if event_type == "INIT_EXP":
                        comm = payload.get("COMM", {})
                        name = comm.get("NAME", "")
                        if name in TASK_NAME_MAP and task_type is None:
                            task_type = TASK_NAME_MAP[name]
                        v_stop_thr = payload.get("V_STOP_THR", 0.01)

                    events.append({"rel_time": rel_time, "event": event_type, "payload": payload})

                    if event_type == "BASE_INFO":
                        base_infos.append({
                            "rel_time": rel_time,
                            "pos_x": payload.get("POS_X", 0),
                            "pos_y": payload.get("POS_Y", 0),
                            "pos_z": payload.get("POS_Z", 0),
                            "speed": payload.get("SPEED", 0),
                            "throttle": payload.get("THROTTLE", 0),
                            "brake": payload.get("BRAKE", 0),
                            "steer": payload.get("STEER_ANGLE", 0),
                        })
        except Exception as e:
            print(f"    WARNING: Error reading {fpath}: {e}")

    # 如果 INIT_EXP 没检测到，用文件名回退
    if task_type is None and fn_task:
        task_type = fn_task

    return {
        "events": events,
        "base_infos": base_infos,
        "task_type": task_type,
        "v_stop_thr": v_stop_thr,
    }


def extract_vr_features(events, base_infos, v_stop_thr):
    """从事件和 BASE_INFO 中提取所有 VR 行为特征"""
    if not base_infos:
        return {}

    speeds = np.array([b["speed"] for b in base_infos])
    throttles = np.array([b["throttle"] for b in base_infos])
    steers = np.array([b["steer"] for b in base_infos])
    times = np.array([b["rel_time"] for b in base_infos])

    features = {}

    # ── Episode 统计 ──
    ep_success = [e for e in events if e["event"] == "EP_SUCCESS"]
    ep_fail = [e for e in events if e["event"] == "EP_FAIL"]
    ep_timeout = [e for e in events if e["event"] == "EP_TIMEOUT"]
    ep_end_events = sorted(ep_success + ep_fail + ep_timeout, key=lambda e: e["rel_time"])

    features["count_ep_success"] = len(ep_success)
    features["count_ep_fail"] = len(ep_fail)
    features["count_ep_timeout"] = len(ep_timeout)
    total_ep = len(ep_success) + len(ep_fail) + len(ep_timeout)
    features["success_rate"] = len(ep_success) / total_ep if total_ep > 0 else 0.0

    # ── Pickup 统计 ──
    pickup_ok = [e for e in events if e["event"] == "PICKUP_OK"]
    pickup_wrong = [e for e in events if e["event"] == "PICKUP_WRONG"]
    features["count_pickup_ok"] = len(pickup_ok)
    features["count_pickup_wrong"] = len(pickup_wrong)
    total_pickup = len(pickup_ok) + len(pickup_wrong)
    features["wrong_pickup_rate"] = len(pickup_wrong) / total_pickup if total_pickup > 0 else 0.0

    # ── 时间 ──
    t_span = float(max(times) - min(times)) if len(times) > 1 else 0.0
    features["duration"] = t_span

    # Episode 时长（EP_START → 对应 EP_END）
    ep_events = sorted(
        [e for e in events if e["event"] in ("EP_START", "EP_SUCCESS", "EP_FAIL", "EP_TIMEOUT")],
        key=lambda e: e["rel_time"]
    )
    episode_durations = []
    success_durations = []
    fail_durations = []
    current_start = None
    for e in ep_events:
        if e["event"] == "EP_START":
            current_start = e["rel_time"]
        elif e["event"] in ("EP_SUCCESS", "EP_FAIL", "EP_TIMEOUT") and current_start is not None:
            dur = e["rel_time"] - current_start
            episode_durations.append(dur)
            if e["event"] == "EP_SUCCESS":
                success_durations.append(dur)
            else:
                fail_durations.append(dur)
            current_start = None

    features["episode_duration_mean"] = float(np.mean(episode_durations)) if episode_durations else 0.0
    features["episode_duration_std"] = float(np.std(episode_durations)) if episode_durations else 0.0
    features["success_time_mean"] = float(np.mean(success_durations)) if success_durations else 0.0
    features["fail_time_mean"] = float(np.mean(fail_durations)) if fail_durations else 0.0

    # ── 地图/导航 ──
    map_ons = sorted([e for e in events if e["event"] == "MAP_ON"], key=lambda e: e["rel_time"])
    map_offs = sorted([e for e in events if e["event"] == "MAP_OFF"], key=lambda e: e["rel_time"])
    features["count_map_on"] = len(map_ons)

    map_view_dur = 0.0
    off_idx = 0
    for mo in map_ons:
        while off_idx < len(map_offs) and map_offs[off_idx]["rel_time"] <= mo["rel_time"]:
            off_idx += 1
        if off_idx < len(map_offs):
            map_view_dur += (map_offs[off_idx]["rel_time"] - mo["rel_time"])

    features["map_view_duration"] = map_view_dur
    features["map_count_per_min"] = (len(map_ons) / t_span * 60) if t_span > 0 else 0.0
    features["map_ratio"] = map_view_dur / t_span if t_span > 0 else 0.0

    # 路径距离
    path_dist = 0.0
    for i in range(1, len(base_infos)):
        dx = base_infos[i]["pos_x"] - base_infos[i-1]["pos_x"]
        dy = base_infos[i]["pos_y"] - base_infos[i-1]["pos_y"]
        dz = base_infos[i]["pos_z"] - base_infos[i-1]["pos_z"]
        path_dist += math.sqrt(dx*dx + dy*dy + dz*dz)
    features["path_distance"] = path_dist

    features["count_zone_enter"] = len([e for e in events if e["event"] == "ZONE_ENTER"])
    features["count_zone_exit"] = len([e for e in events if e["event"] == "ZONE_EXIT"])

    # ── 驾驶行为 ──
    features["speed_mean"] = float(np.mean(speeds)) if len(speeds) > 0 else 0.0
    features["speed_std"] = float(np.std(speeds)) if len(speeds) > 0 else 0.0
    features["speed_max"] = float(np.max(speeds)) if len(speeds) > 0 else 0.0
    features["stationary_ratio"] = float(np.sum(speeds < v_stop_thr) / len(speeds)) if len(speeds) > 0 else 0.0

    features["count_brake_on"] = len([e for e in events if e["event"] == "BRAKE_ON"])
    features["brake_count_per_min"] = (features["count_brake_on"] / t_span * 60) if t_span > 0 else 0.0

    features["throttle_mean"] = float(np.mean(throttles)) if len(throttles) > 0 else 0.0

    abs_steers = np.abs(steers)
    features["abs_steer_mean"] = float(np.mean(abs_steers)) if len(abs_steers) > 0 else 0.0
    features["abs_steer_std"] = float(np.std(abs_steers)) if len(abs_steers) > 0 else 0.0

    # ── 注意力/停车 ──
    features["count_att_start"] = len([e for e in events if e["event"] == "ATT_START"])
    features["count_att_reset"] = len([e for e in events if e["event"] == "ATT_RESET"])
    features["count_stop_start"] = len([e for e in events if e["event"] == "STOP_START"])

    # stop_duration: STOP_START → STOP_END
    stop_starts = sorted([e for e in events if e["event"] == "STOP_START"], key=lambda e: e["rel_time"])
    stop_ends = sorted([e for e in events if e["event"] == "STOP_END"], key=lambda e: e["rel_time"])
    stop_dur = 0.0
    end_idx = 0
    for ss in stop_starts:
        while end_idx < len(stop_ends) and stop_ends[end_idx]["rel_time"] <= ss["rel_time"]:
            end_idx += 1
        if end_idx < len(stop_ends):
            stop_dur += (stop_ends[end_idx]["rel_time"] - ss["rel_time"])
    features["stop_duration"] = stop_dur
    features["stop_ratio"] = stop_dur / t_span if t_span > 0 else 0.0

    return features


# ─── 量表 Excel ────────────────────────────────────────

def parse_cdr(raw):
    """解析 CDR: 'global-0.5_SB-1' → (0.5, 1)"""
    if raw is None:
        return np.nan, np.nan
    raw = str(raw).strip()
    m = re.match(r"global-([\d.]+)_SB-([\d.]+)", raw.replace(" ", ""))
    if m:
        return float(m.group(1)), float(m.group(2))
    try:
        return float(raw), np.nan
    except ValueError:
        pass
    return np.nan, np.nan


def extract_all_scales(scale_dir):
    """提取所有被试的量表总分"""
    scale_data = {}
    for fname in sorted(os.listdir(scale_dir)):
        if fname.startswith("~$") or not fname.endswith(".xlsx"):
            continue
        fpath = os.path.join(scale_dir, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception:
            continue
        ws = wb[wb.sheetnames[-1]]

        raw_id = str(ws.cell(row=1, column=1).value or "").strip()
        sid = raw_id.replace("_____ATHENA", "").replace("_____", "").strip()
        sid = sid.replace("ATHENA", "ATH")
        if not sid.startswith("ATH"):
            sid = "ATH" + sid

        def safe_float(v):
            try:
                return float(v) if v is not None else np.nan
            except (ValueError, TypeError):
                return np.nan

        mmse = safe_float(ws.cell(row=4, column=2).value)
        moca = safe_float(ws.cell(row=5, column=2).value)
        cdr_global, cdr_sb = parse_cdr(ws.cell(row=6, column=2).value)
        his = safe_float(ws.cell(row=7, column=2).value)

        scale_data[sid] = {
            "MMSE": mmse, "MOCA": moca,
            "CDR_global": cdr_global, "CDR_SB": cdr_sb, "HIS": his,
        }
    return scale_data


# ─── 总体 & 差异特征 ──────────────────────────────────

def compute_overall_features(grid4_feats, grid9_feats):
    """合并 grid4 + grid9 计算 overall 特征"""
    overall = {}
    if not grid4_feats and not grid9_feats:
        return overall

    # 合并事件计数的字段
    count_fields = ["count_ep_success", "count_ep_fail", "count_ep_timeout",
                    "count_pickup_ok", "count_pickup_wrong"]
    for field in count_fields:
        total = 0
        if grid4_feats:
            total += grid4_feats.get(field, 0)
        if grid9_feats:
            total += grid9_feats.get(field, 0)
        overall[f"_total_{field}"] = total

    total_ep = (overall.get("_total_count_ep_success", 0) +
                overall.get("_total_count_ep_fail", 0) +
                overall.get("_total_count_ep_timeout", 0))
    overall["success_rate"] = overall.get("_total_count_ep_success", 0) / total_ep if total_ep > 0 else 0.0

    total_ok = overall.get("_total_count_pickup_ok", 0)
    total_wrong = overall.get("_total_count_pickup_wrong", 0)
    total_pu = total_ok + total_wrong
    overall["wrong_pickup_rate"] = total_wrong / total_pu if total_pu > 0 else 0.0

    overall["duration"] = (grid4_feats.get("duration", 0) if grid4_feats else 0) + \
                          (grid9_feats.get("duration", 0) if grid9_feats else 0)

    total_map_view = (grid4_feats.get("map_view_duration", 0) if grid4_feats else 0) + \
                     (grid9_feats.get("map_view_duration", 0) if grid9_feats else 0)
    overall["map_ratio"] = total_map_view / overall["duration"] if overall["duration"] > 0 else 0.0

    overall["path_distance"] = (grid4_feats.get("path_distance", 0) if grid4_feats else 0) + \
                               (grid9_feats.get("path_distance", 0) if grid9_feats else 0)

    # 加权平均速度
    d4 = grid4_feats.get("duration", 0) if grid4_feats else 0
    d9 = grid9_feats.get("duration", 0) if grid9_feats else 0
    total_dur = d4 + d9
    if total_dur > 0:
        overall["speed_mean"] = ((grid4_feats.get("speed_mean", 0) if grid4_feats else 0) * d4 +
                                 (grid9_feats.get("speed_mean", 0) if grid9_feats else 0) * d9) / total_dur
        overall["speed_std"] = math.sqrt(
            ((grid4_feats.get("speed_std", 0) if grid4_feats else 0)**2 * d4 +
             (grid9_feats.get("speed_std", 0) if grid9_feats else 0)**2 * d9) / total_dur
        ) if total_dur > 0 else 0.0
    else:
        overall["speed_mean"] = 0.0
        overall["speed_std"] = 0.0

    overall["stationary_ratio"] = (
        (grid4_feats.get("stationary_ratio", 0) * d4 if grid4_feats else 0) +
        (grid9_feats.get("stationary_ratio", 0) * d9 if grid9_feats else 0)
    ) / total_dur if total_dur > 0 else 0.0

    total_stop = (grid4_feats.get("stop_duration", 0) if grid4_feats else 0) + \
                 (grid9_feats.get("stop_duration", 0) if grid9_feats else 0)
    overall["stop_ratio"] = total_stop / total_dur if total_dur > 0 else 0.0

    # 清理中间字段
    for field in count_fields:
        overall.pop(f"_total_{field}", None)

    return overall


def compute_diff_features(grid4_feats, grid9_feats):
    """grid9 - grid4 差异特征"""
    diff = {}
    for key in ["success_rate", "wrong_pickup_rate", "map_ratio", "stop_ratio", "path_distance"]:
        g9 = grid9_feats.get(key, np.nan) if grid9_feats else np.nan
        g4 = grid4_feats.get(key, np.nan) if grid4_feats else np.nan
        try:
            diff[key] = float(g9) - float(g4)
        except (TypeError, ValueError):
            diff[key] = np.nan
    return diff


# ─── 主流程 ─────────────────────────────────────────────

def val_to_str(v):
    """将值转换为 CSV 友好的字符串，NaN → 空"""
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return ""
    return v


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ─── Step 1: VR 日志 ───
    print("=" * 60)
    print("Step 1/4: Parsing VR logs...")

    vr_features_raw = {}  # subject_id -> {"grid4": {...}, "grid9": {...}}

    for folder_name in sorted(os.listdir(VR_LOG_DIR)):
        folder_path = os.path.join(VR_LOG_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue

        subject_id = folder_name.replace("VR", "").strip()

        # 收集所有 .log 文件，按文件名排序
        all_logs = sorted([
            os.path.join(folder_path, f)
            for f in os.listdir(folder_path) if f.endswith(".log")
        ])

        if not all_logs:
            continue

        # 按任务类型分组
        grid4_files = []
        grid9_files = []
        unknown_files = []

        for fpath in all_logs:
            fname = os.path.basename(fpath)
            # 首先从文件名尝试
            fn_task = detect_task_type_from_filename(fname)
            if fn_task == "grid4":
                grid4_files.append(fpath)
                continue
            elif fn_task == "grid9":
                grid9_files.append(fpath)
                continue

            # 从 INIT_EXP 读取
            init_task = read_init_exp_task_type(fpath)
            if init_task == "grid4":
                grid4_files.append(fpath)
            elif init_task == "grid9":
                grid9_files.append(fpath)
            else:
                unknown_files.append(fpath)

        # 处理无法识别的文件：继承前一个已知任务类型
        if unknown_files:
            # 尝试从同一目录中已知类型文件推断
            for uf in unknown_files:
                # 找时间戳最接近的前一个已知文件
                uf_ts = os.path.basename(uf).split("_")[0]
                last_type = None
                for fpath in sorted(all_logs):
                    if fpath >= uf:
                        break
                    fname = os.path.basename(fpath)
                    fn_task = detect_task_type_from_filename(fname)
                    if fn_task is None:
                        fn_task = read_init_exp_task_type(fpath)
                    if fn_task:
                        last_type = fn_task
                if last_type == "grid4":
                    grid4_files.append(uf)
                elif last_type == "grid9":
                    grid9_files.append(uf)
                # 如果仍无法确定，跳过

        # 提取特征
        result = {"grid4": None, "grid9": None}

        if grid4_files:
            parsed = parse_log_events(grid4_files)
            if parsed["base_infos"]:
                feats = extract_vr_features(parsed["events"], parsed["base_infos"], parsed["v_stop_thr"])
                result["grid4"] = feats

        if grid9_files:
            parsed = parse_log_events(grid9_files)
            if parsed["base_infos"]:
                feats = extract_vr_features(parsed["events"], parsed["base_infos"], parsed["v_stop_thr"])
                result["grid9"] = feats

        vr_features_raw[subject_id] = result

    print(f"  Subjects with VR data: {len(vr_features_raw)}")
    grid4_count = sum(1 for v in vr_features_raw.values() if v["grid4"])
    grid9_count = sum(1 for v in vr_features_raw.values() if v["grid9"])
    print(f"    grid4: {grid4_count}")
    print(f"    grid9: {grid9_count}")

    # ─── Step 2: 量表 ───
    print("\nStep 2/4: Extracting scale scores...")
    scale_data = extract_all_scales(SCALE_DIR)
    print(f"  Subjects with scale: {len(scale_data)}")

    # ─── Step 3: 笔迹特征 ───
    print("\nStep 3/4: Reading drawing features...")
    drawing_data = {}
    drawing_cols_order = []
    with open(DRAWING_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        drawing_cols_order = [c for c in reader.fieldnames if c != "subject_id"]
        for row in reader:
            sid = row["subject_id"].strip()
            if not sid.startswith("ATH"):
                sid = "ATH" + sid
            drawing_data[sid] = {k: row[k] for k in drawing_cols_order}
    print(f"  Subjects with drawing: {len(drawing_data)}")

    # ─── Step 4: 合并 ───
    print("\nStep 4/4: Merging all data sources...")

    all_subjects = sorted(set(
        list(vr_features_raw.keys()) +
        list(scale_data.keys()) +
        list(drawing_data.keys())
    ))
    print(f"  Total unique subjects: {len(all_subjects)}")

    # 列定义
    SCALE_COLS = ["MMSE", "MOCA", "CDR_global", "CDR_SB", "HIS"]

    VR_FIELDS = [
        "success_rate", "count_ep_success", "count_ep_fail", "count_ep_timeout",
        "count_pickup_ok", "count_pickup_wrong", "wrong_pickup_rate",
        "duration", "episode_duration_mean", "episode_duration_std",
        "success_time_mean", "fail_time_mean",
        "count_map_on", "map_view_duration", "map_count_per_min", "map_ratio",
        "path_distance", "count_zone_enter", "count_zone_exit",
        "speed_mean", "speed_std", "speed_max", "stationary_ratio",
        "count_brake_on", "brake_count_per_min", "throttle_mean",
        "abs_steer_mean", "abs_steer_std",
        "count_att_start", "count_att_reset",
        "count_stop_start", "stop_duration", "stop_ratio",
    ]
    GRID4_COLS = [f"grid4_{f}" for f in VR_FIELDS]
    GRID9_COLS = [f"grid9_{f}" for f in VR_FIELDS]

    OVERALL_FIELDS = [
        "success_rate", "wrong_pickup_rate", "duration", "map_ratio",
        "path_distance", "speed_mean", "speed_std", "stationary_ratio", "stop_ratio",
    ]
    OVERALL_COLS = [f"overall_{f}" for f in OVERALL_FIELDS]

    DIFF_FIELDS = [
        "success_rate", "wrong_pickup_rate", "map_ratio", "stop_ratio", "path_distance",
    ]
    DIFF_COLS = [f"diff_{f}" for f in DIFF_FIELDS]

    ALL_COLS = ["subject_id"] + SCALE_COLS + GRID4_COLS + GRID9_COLS + \
               OVERALL_COLS + DIFF_COLS + drawing_cols_order

    # 构建输出
    merged_rows = []
    vr_output_rows = []

    for sid in all_subjects:
        row = {"subject_id": sid}

        # 量表
        scales = scale_data.get(sid, {})
        for col in SCALE_COLS:
            row[col] = val_to_str(scales.get(col, ""))

        # grid4 / grid9
        g4_feats = vr_features_raw.get(sid, {}).get("grid4") or {}
        g9_feats = vr_features_raw.get(sid, {}).get("grid9") or {}

        for col in GRID4_COLS:
            key = col.replace("grid4_", "")
            row[col] = val_to_str(g4_feats.get(key, ""))
        for col in GRID9_COLS:
            key = col.replace("grid9_", "")
            row[col] = val_to_str(g9_feats.get(key, ""))

        # overall
        overall = compute_overall_features(g4_feats, g9_feats)
        for col in OVERALL_COLS:
            key = col.replace("overall_", "")
            row[col] = val_to_str(overall.get(key, ""))

        # diff
        diff = compute_diff_features(g4_feats, g9_feats)
        for col in DIFF_COLS:
            key = col.replace("diff_", "")
            row[col] = val_to_str(diff.get(key, ""))

        # drawing
        draw = drawing_data.get(sid, {})
        for col in drawing_cols_order:
            row[col] = val_to_str(draw.get(col, ""))

        merged_rows.append(row)

        # VR 中间表行
        vr_row = {"subject_id": sid}
        for col in GRID4_COLS:
            key = col.replace("grid4_", "")
            vr_row[col] = val_to_str(g4_feats.get(key, ""))
        for col in GRID9_COLS:
            key = col.replace("grid9_", "")
            vr_row[col] = val_to_str(g9_feats.get(key, ""))
        for col in OVERALL_COLS:
            key = col.replace("overall_", "")
            vr_row[col] = val_to_str(overall.get(key, ""))
        for col in DIFF_COLS:
            key = col.replace("diff_", "")
            vr_row[col] = val_to_str(diff.get(key, ""))
        vr_output_rows.append(vr_row)

    # ── 写入文件 ──
    # merged_dataset.csv
    output_csv = os.path.join(OUTPUT_DIR, "merged_dataset.csv")
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_COLS)
        writer.writeheader()
        for row in merged_rows:
            writer.writerow(row)
    print(f"\n=> {output_csv}")
    print(f"   {len(merged_rows)} subjects x {len(ALL_COLS)} columns")

    # scale_scores.csv
    scale_csv = os.path.join(OUTPUT_DIR, "scale_scores.csv")
    with open(scale_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["subject_id"] + SCALE_COLS)
        writer.writeheader()
        for sid in sorted(scale_data.keys()):
            r = {"subject_id": sid}
            for col in SCALE_COLS:
                r[col] = val_to_str(scale_data[sid].get(col, ""))
            writer.writerow(r)
    print(f"=> {scale_csv}")

    # vr_features.csv
    vr_csv = os.path.join(OUTPUT_DIR, "vr_features.csv")
    vr_all_cols = ["subject_id"] + GRID4_COLS + GRID9_COLS + OVERALL_COLS + DIFF_COLS
    with open(vr_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=vr_all_cols)
        writer.writeheader()
        for row in vr_output_rows:
            writer.writerow(row)
    print(f"=> {vr_csv}")

    # ── 数据完整性统计 ──
    n_scale = sum(1 for r in merged_rows if r["MMSE"] != "")
    n_grid4 = sum(1 for r in merged_rows if r["grid4_duration"] != "")
    n_grid9 = sum(1 for r in merged_rows if r["grid9_duration"] != "")
    n_drawing = sum(1 for r in merged_rows
                    if drawing_cols_order and r.get(drawing_cols_order[0], "") != "")
    n_overall = sum(1 for r in merged_rows if r["overall_duration"] != "")
    n_diff = sum(1 for r in merged_rows if r["diff_success_rate"] != "")

    n_full = sum(1 for r in merged_rows
                 if r["MMSE"] != "" and r["grid4_duration"] != "" and r["grid9_duration"] != "")

    print(f"\n  Data completeness ({len(merged_rows)} subjects):")
    print(f"    量表 (MMSE/MOCA/CDR/HIS): {n_scale}")
    print(f"    drawing_* 笔迹特征:       {n_drawing}")
    print(f"    grid4_* VR四宫格:         {n_grid4}")
    print(f"    grid9_* VR九宫格:         {n_grid9}")
    print(f"    overall_* 总体VR:         {n_overall}")
    print(f"    diff_* 复杂度差异:        {n_diff}")
    print(f"    完整(量表+grid4+grid9):   {n_full}")

    # ── 完整性报告 ──
    status_txt = os.path.join(OUTPUT_DIR, "merged_dataset_status.txt")
    with open(status_txt, "w", encoding="utf-8") as f:
        f.write("merged_dataset.csv 数据完整性报告\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"生成时间: 2026-06-08\n")
        f.write(f"被试数: {len(merged_rows)}\n")
        f.write(f"总列数: {len(ALL_COLS)}\n\n")
        f.write(f"列分布:\n")
        f.write(f"  subject_id:               1\n")
        f.write(f"  量表总分 (MMSE等):        {len(SCALE_COLS)}\n")
        f.write(f"  grid4_* VR四宫格:         {len(GRID4_COLS)}\n")
        f.write(f"  grid9_* VR九宫格:         {len(GRID9_COLS)}\n")
        f.write(f"  overall_* 总体VR:         {len(OVERALL_COLS)}\n")
        f.write(f"  diff_* 复杂度差异:        {len(DIFF_COLS)}\n")
        f.write(f"  drawing_* 笔迹特征:       {len(drawing_cols_order)}\n\n")
        f.write(f"数据填充情况:\n")
        f.write(f"  量表:      {n_scale}/{len(merged_rows)}\n")
        f.write(f"  grid4:     {n_grid4}/{len(merged_rows)}\n")
        f.write(f"  grid9:     {n_grid9}/{len(merged_rows)}\n")
        f.write(f"  overall:   {n_overall}/{len(merged_rows)}\n")
        f.write(f"  diff:      {n_diff}/{len(merged_rows)}\n")
        f.write(f"  drawing:   {n_drawing}/{len(merged_rows)}\n")
        f.write(f"  全数据:    {n_full}/{len(merged_rows)}\n")
    print(f"=> {status_txt}")

    print("\nDone! merged_dataset.csv is ready.")


if __name__ == "__main__":
    main()
